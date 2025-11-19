import pyautogui
import time
import subprocess
import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
import shutil

# Configuração das lojas
lojas = {
    "Mooca": {"login": "ljmooca", "senha": "luc1549", "setinhas": 9, "tecla_inicial": "m", "controle_caixa": (324, 504)},
    "Lorena": {"login": "LJALORENA", "senha": "luc1549", "setinhas": 8, "tecla_inicial": "m", "controle_caixa": (332, 536)},
    "Alto de Pinheiros": {"login": "LJALTOPINHEIROS", "senha": "AVATIM231", "setinhas": 6, "tecla_inicial": "m", "controle_caixa": (359, 532)},
    "Ibirapuera": {"login": "LJIBIRAPUERA", "senha": "AVATIM123", "setinhas": 7, "tecla_inicial": "m", "controle_caixa": (361, 535)},
    "Perdizes": {"login": "LJF2PERDIZES", "senha": "avatim123", "setinhas": 10, "tecla_inicial": "f", "controle_caixa": (398, 501)},
    "Leopoldina": {"login": "LJVLEOPOLDINA", "senha": "AVATIM123", "setinhas": 12, "tecla_inicial": "f", "controle_caixa": (361, 504)},
    "Santana": {"login": "LJF2SANTANA", "senha": "AVATIM123", "setinhas": 11, "tecla_inicial": "f", "controle_caixa": (365, 506)}
}

chrome_path = "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
exatto_url = "https://portalfranquias.avatim.com.br/"
caminho_destino = r"C:\Users\PERFIL\Documents\Controle de Caixa"
os.makedirs(caminho_destino, exist_ok=True)
planilha_f360 = os.path.join(caminho_destino, "Planilha F360.xlsx")
caminho_downloads = r"C:\Users\PERFIL\Downloads"

# Calcula datas
hoje = datetime.today()
ontem = hoje - timedelta(days=1)
if hoje.weekday() == 0:
    sexta = hoje - timedelta(days=3)
    domingo = hoje - timedelta(days=1)
    DATA_INICIAL = sexta.strftime("%d/%m/%Y")
    DATA_FINAL = domingo.strftime("%d/%m/%Y")
else:
    DATA_INICIAL = ontem.strftime("%d/%m/%Y")
    DATA_FINAL = ontem.strftime("%d/%m/%Y")

print(f"📅 Período: {DATA_INICIAL} até {DATA_FINAL}")

# -------------------------------
# Função para processar loja (mantida do seu código original)
# -------------------------------
def processar_loja(nome_loja, login, senha, setinhas, tecla_inicial, controle_caixa):
    print(f"\n🔄 Processando loja: {nome_loja}")

    subprocess.Popen([chrome_path, exatto_url])
    time.sleep(5)

    pyautogui.click(x=697, y=423)
    pyautogui.hotkey("ctrl", "a"); pyautogui.press("backspace")
    pyautogui.write(login)

    pyautogui.click(x=473, y=490)
    pyautogui.hotkey("ctrl", "a"); pyautogui.press("backspace")
    pyautogui.write(senha)

    pyautogui.click(x=610, y=549)
    time.sleep(1)
    pyautogui.press(tecla_inicial)
    time.sleep(1)
    for _ in range(setinhas):
        pyautogui.press("down")
        time.sleep(0.2)
    pyautogui.press("enter")

    pyautogui.click(x=1035, y=549)
    time.sleep(3)

    pyautogui.click(x=276, y=253)
    time.sleep(1)
    x_cc, y_cc = controle_caixa
    pyautogui.click(x=x_cc, y=y_cc)
    time.sleep(2)

    pyautogui.click(x=200, y=458)
    pyautogui.hotkey("ctrl", "a"); pyautogui.press("backspace")
    pyautogui.write(DATA_INICIAL)

    pyautogui.click(x=198, y=483)
    pyautogui.hotkey("ctrl", "a"); pyautogui.press("backspace")
    pyautogui.write(DATA_FINAL)

    pyautogui.click(x=1000, y=560)
    time.sleep(3)

    pyautogui.moveTo(x=1300, y=600)
    time.sleep(1)
    for _ in range(2):
        pyautogui.scroll(-1000)
        time.sleep(1)

    for f in os.listdir(caminho_downloads):
        if f.startswith("Result") and f.endswith(".xls"):
            os.remove(os.path.join(caminho_downloads, f))

    pyautogui.click(x=1052, y=482)
    time.sleep(7)

    arquivo_download = os.path.join(caminho_downloads, "Result.xls")
    arquivo_destino = os.path.join(caminho_destino, "Result.xls")
    if os.path.exists(arquivo_download):
        shutil.move(arquivo_download, arquivo_destino)

    tables = pd.read_html(arquivo_destino, header=0)
    df = tables[0]
    df = df.drop(columns=["ABERTURA", "FECHAMENTO"], errors="ignore")

    df_f360 = df[["DATA", "VALOR", "FORMA DE PAGAMENTO", "CLIENTE/FORNECEDOR", "OPERADOR", "NOTA FISCAL"]]
    df_f360 = df_f360.rename(columns={
        "DATA": "Data Venda",
        "VALOR": "Valor Bruto",
        "FORMA DE PAGAMENTO": "Forma de Pagamento",
        "CLIENTE/FORNECEDOR": "Nome do Cliente",
        "OPERADOR": "Operador",
        "NOTA FISCAL": "Nota Fiscal"
    })

    book = load_workbook(planilha_f360)
    sheet = book["Modelo de Importação de GFF-PDV"]

    max_row = sheet.max_row
    for row in range(3, max_row+1):
        for col in range(1, sheet.max_column+1):
            sheet.cell(row=row, column=col).value = None

    start_row = 3
    for i, row in df_f360.iterrows():
        sheet.cell(row=start_row+i, column=1, value=row["Data Venda"])
        sheet.cell(row=start_row+i, column=2, value=row["Valor Bruto"])
        sheet.cell(row=start_row+i, column=3, value=row["Forma de Pagamento"])
        sheet.cell(row=start_row+i, column=12, value=row["Nome do Cliente"])
        sheet.cell(row=start_row+i, column=15, value=row["Operador"])
        sheet.cell(row=start_row+i, column=16, value=row["Nota Fiscal"])

    book.save(planilha_f360)

    data_str = (datetime.today() - timedelta(days=1)).strftime("%d-%m-%Y")
    copia_planilha = os.path.join(caminho_destino, f"{nome_loja}_{data_str}.xls")
    shutil.copy(planilha_f360, copia_planilha)

    print(f"✅ Loja {nome_loja} processada. Cópia criada: {copia_planilha}")
    return copia_planilha

# -------------------------------
# Função para upload no F360 (nova)
# -------------------------------
def upload_f360(lista_arquivos):
    subprocess.Popen([chrome_path, "https://financas.f360.com.br/"])
    time.sleep(5)

    pyautogui.click(x=159, y=583)
    time.sleep(1)

    pyautogui.click(x=1271, y=340)
    pyautogui.write("pedro.noronha@manjos.com.br")

    pyautogui.click(x=895, y=422)
    pyautogui.write("Adm2025!")

    pyautogui.click(x=900, y=535)
    time.sleep(3)
    pyautogui.click(x=900, y=535)
    time.sleep(10)

    pyautogui.moveTo(x=159, y=583)
    time.sleep(1)
    pyautogui.scroll(-120)
    time.sleep(3)
    pyautogui.click(x=159, y=583)
    time.sleep(3)

    pyautogui.click(x=387, y=346)
    time.sleep(3)

    for arquivo in lista_arquivos:
        pyautogui.write(arquivo)
        pyautogui.press("enter")
        time.sleep(2)

    print("📤 Upload concluído para todos os arquivos.")

# -------------------------------
# Execução principal
# -------------------------------
if __name__ == "__main__":
    arquivos_gerados = []
    for nome_loja, config in lojas.items():
        copia_planilha = processar_loja(
            nome_loja,
            config["login"],
            config["senha"],
            config["setinhas"],
            config["tecla_inicial"],
            config["controle_caixa"]
        )
        arquivos_gerados.append(copia_planilha)

    # 🚀 Agora faz o upload no F360
    upload_f360(arquivos_gerados)
